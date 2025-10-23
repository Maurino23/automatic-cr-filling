import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import time

st.set_page_config(page_title="Automated CR Filling", layout="wide", initial_sidebar_state="expanded")

# Initialize session state for additional input files
if 'additional_files' not in st.session_state:
    st.session_state.additional_files = {}

# Sidebar untuk informasi dan pengaturan
with st.sidebar:
    st.image("E:\RINO\DATA ANALYST\CR Automatic Filling\Logo App.png", use_container_width=True)
    st.title("ℹ️ Informasi Aplikasi")
    st.info("""
    **Automated CR Filling App v2.5**
    
    Aplikasi ini membantu Anda mengisi Collective Roster secara otomatis dengan fitur:
    - ✅ Multi-tanggal (Range & Manual)
    - ✅ File input berbeda per tanggal
    - ✅ Validasi data otomatis
    - 📊 Statistik lengkap
    - 🎨 Format Excel profesional
    - 📝 Log aktivitas
    """)
    
    st.divider()
    st.subheader("⚙️ Pengaturan")
    apply_formatting = st.checkbox("Terapkan Format Excel", value=True, help="Format warna dan style pada hasil Excel")
    show_stats = st.checkbox("Tampilkan Statistik", value=True, help="Menampilkan statistik detail hasil pemrosesan")
    
    st.divider()
    st.caption("© 2025 Automated CR Filling")

# Header utama
st.title("📋 Automated CR Filling App")
st.caption("Versi 2.5 - Multi-Date Processing")

# Layout upload file
st.divider()
col1, col2 = st.columns(2)

with col1:
    st.header("📂 Upload Template")
    template_file = st.file_uploader(
        "Upload Template File (.xlsx)", 
        type=["xlsx"], 
        key="template",
        help="File template yang akan diisi"
    )
    if template_file:
        file_details = {
            "Nama File": template_file.name,
            "Ukuran": f"{template_file.size / 1024:.2f} KB",
            "Tipe": template_file.type
        }
        with st.expander("📄 Detail File Template"):
            for key, value in file_details.items():
                st.text(f"{key}: {value}")

with col2:
    st.header("📥 Upload Input Utama")
    input_file = st.file_uploader(
        "Upload File Input Utama (.xlsx)", 
        type=["xlsx"], 
        key="input",
        help="File data yang akan digunakan sebagai default untuk semua tanggal"
    )
    if input_file:
        file_details = {
            "Nama File": input_file.name,
            "Ukuran": f"{input_file.size / 1024:.2f} KB",
            "Tipe": input_file.type
        }
        with st.expander("📄 Detail File Input"):
            for key, value in file_details.items():
                st.text(f"{key}: {value}")

# Parameter proses - Multi Date
st.divider()
st.header("🗓️ Parameter Tanggal")

# Mode selection
date_mode = st.radio(
    "Pilih Mode Input Tanggal:",
    ["📅 Range Tanggal", "🎯 Pilih Manual"],
    horizontal=True,
    help="Pilih apakah ingin menggunakan range atau memilih tanggal secara manual"
)

selected_dates = []

if date_mode == "📅 Range Tanggal":
    col_range1, col_range2 = st.columns(2)
    with col_range1:
        start_date = st.number_input("Dari Tanggal", min_value=1, max_value=31, value=1)
    with col_range2:
        end_date = st.number_input("Sampai Tanggal", min_value=1, max_value=31, value=5)
    
    if start_date > end_date:
        st.error("❌ Tanggal awal tidak boleh lebih besar dari tanggal akhir!")
    else:
        selected_dates = list(range(start_date, end_date + 1))
        st.success(f"✅ Akan memproses {len(selected_dates)} tanggal: {', '.join(map(str, selected_dates))}")
else:
    selected_dates = st.multiselect(
        "Pilih Tanggal yang Ingin Diproses:",
        options=list(range(1, 32)),
        default=[1, 2, 3],
        help="Pilih satu atau lebih tanggal yang ingin diproses"
    )
    if selected_dates:
        st.info(f"ℹ️ Terpilih {len(selected_dates)} tanggal: {', '.join(map(str, sorted(selected_dates)))}")

# Additional files per date (optional)
if selected_dates:
    st.divider()
    with st.expander("📎 Upload File Input Berbeda per Tanggal (Opsional)", expanded=False):
        st.caption("Jika tidak di-upload, akan menggunakan file input utama untuk tanggal tersebut")
        
        cols_per_row = 3
        for i in range(0, len(selected_dates), cols_per_row):
            cols = st.columns(cols_per_row)
            for j, col in enumerate(cols):
                if i + j < len(selected_dates):
                    date = selected_dates[i + j]
                    with col:
                        additional_file = st.file_uploader(
                            f"Tanggal {date}",
                            type=["xlsx"],
                            key=f"additional_file_{date}",
                            help=f"File input khusus untuk tanggal {date}"
                        )
                        if additional_file:
                            st.session_state.additional_files[date] = additional_file
                            st.success(f"✓ File untuk tgl {date}")
                        elif date in st.session_state.additional_files:
                            del st.session_state.additional_files[date]

# Output filename
st.divider()
output_filename = st.text_input(
    "Nama File Hasil (tanpa ekstensi)", 
    value=f"CR_FILLED_MULTI_{datetime.now().strftime('%Y%m%d')}",
    help="Nama file output yang akan diunduh"
)

# Preview data
st.divider()
st.header("🔍 Preview & Validasi Data")

template_df = None
input_data_df = None
validation_errors = []

# Tab untuk preview
tab1, tab2, tab3 = st.tabs(["📑 Template File", "📄 Input Utama", "⚠️ Validasi"])

with tab1:
    if template_file:
        try:
            template_df = pd.read_excel(template_file, sheet_name=0, header=1)
            
            col_info1, col_info2, col_info3 = st.columns(3)
            with col_info1:
                st.metric("Total Baris", len(template_df))
            with col_info2:
                st.metric("Total Kolom", len(template_df.columns))
            with col_info3:
                if "Crew ID" in template_df.columns:
                    st.metric("Crew ID Unik", template_df["Crew ID"].nunique())
            
            st.dataframe(template_df, use_container_width=True, height=400)
                
        except Exception as e:
            st.error(f"❌ Gagal membaca file template: {e}")
            validation_errors.append(f"Error template: {e}")
    else:
        st.info("👆 Silakan upload file template terlebih dahulu")

with tab2:
    if input_file:
        try:
            input_data_df = pd.read_excel(input_file, sheet_name=0, header=1)
            
            col_info1, col_info2, col_info3 = st.columns(3)
            with col_info1:
                st.metric("Total Baris", len(input_data_df))
            with col_info2:
                st.metric("Total Kolom", len(input_data_df.columns))
            with col_info3:
                if "Crew ID" in input_data_df.columns:
                    st.metric("Crew ID Unik", input_data_df["Crew ID"].nunique())
            
            st.dataframe(input_data_df, use_container_width=True, height=400)
                
        except Exception as e:
            st.error(f"❌ Gagal membaca file input: {e}")
            validation_errors.append(f"Error input: {e}")
    else:
        st.info("👆 Silakan upload file input terlebih dahulu")

with tab3:
    if template_df is not None and input_data_df is not None:
        # Validasi kolom
        if "Crew ID" not in template_df.columns:
            validation_errors.append("❌ Template tidak memiliki kolom 'Crew ID'")
        if "Crew ID" not in input_data_df.columns:
            validation_errors.append("❌ Input tidak memiliki kolom 'Crew ID'")
        
        # Validasi tanggal yang dipilih
        if not selected_dates:
            validation_errors.append("❌ Belum ada tanggal yang dipilih untuk diproses")
        else:
            missing_date_cols = []
            for date in selected_dates:
                date_col = None
                for col in template_df.columns:
                    if str(date) in str(col):
                        date_col = col
                        break
                if date_col is None:
                    missing_date_cols.append(date)
            
            if missing_date_cols:
                validation_errors.append(f"❌ Kolom tanggal tidak ditemukan di template untuk: {', '.join(map(str, missing_date_cols))}")
        
        # Tampilkan hasil validasi
        if validation_errors:
            st.error("**Ditemukan masalah validasi:**")
            for error in validation_errors:
                st.write(error)
        else:
            st.success("✅ Semua validasi berhasil! Siap untuk diproses.")
            
            # Tampilkan perbandingan Crew ID
            if "Crew ID" in template_df.columns and "Crew ID" in input_data_df.columns:
                template_ids = set(template_df["Crew ID"].astype(str).str.strip())
                input_ids = set(input_data_df["Crew ID"].astype(str).str.strip())
                
                col_comp1, col_comp2, col_comp3 = st.columns(3)
                with col_comp1:
                    st.metric("ID di Template", len(template_ids))
                with col_comp2:
                    st.metric("ID di Input", len(input_ids))
                with col_comp3:
                    matched = len(template_ids.intersection(input_ids))
                    st.metric("ID yang Cocok", matched)
                
                # ID yang tidak cocok
                missing_in_input = template_ids - input_ids
                missing_in_template = input_ids - template_ids
                
                if missing_in_input:
                    with st.expander(f"⚠️ ID di Template tapi tidak di Input ({len(missing_in_input)})"):
                        st.write(list(missing_in_input)[:20])
                        if len(missing_in_input) > 20:
                            st.caption(f"... dan {len(missing_in_input) - 20} lainnya")
                
                if missing_in_template:
                    with st.expander(f"ℹ️ ID di Input tapi tidak di Template ({len(missing_in_template)})"):
                        st.write(list(missing_in_template)[:20])
                        if len(missing_in_template) > 20:
                            st.caption(f"... dan {len(missing_in_template) - 20} lainnya")
            
            # Info file tambahan
            if st.session_state.additional_files:
                st.divider()
                st.subheader("📎 File Input Tambahan")
                additional_info = []
                for date, file in st.session_state.additional_files.items():
                    additional_info.append(f"Tanggal {date}: {file.name}")
                st.info("File khusus akan digunakan untuk:\n" + "\n".join(f"• {info}" for info in additional_info))
    else:
        st.info("👆 Upload kedua file untuk melihat hasil validasi")

# Proses data
st.divider()
st.header("⚙️ Proses Data")

process_disabled = (template_df is None or input_data_df is None or 
                   len(validation_errors) > 0 or not selected_dates)

if st.button("🚀 Proses Semua Tanggal", type="primary", use_container_width=True, disabled=process_disabled):
    
    start_time = time.time()
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    try:
        status_text.text("📊 Memulai pemrosesan multi-tanggal...")
        progress_bar.progress(5)
        
        # Statistik keseluruhan
        total_dates = len(selected_dates)
        overall_stats = {
            'total_rows': len(template_df),
            'dates_processed': 0,
            'total_matched': 0,
            'total_not_matched': 0,
            'total_empty': 0,
            'date_details': {}
        }
        
        # Proses setiap tanggal
        for idx, date in enumerate(selected_dates):
            status_text.text(f"🔄 Memproses tanggal {date} ({idx + 1}/{total_dates})...")
            
            # Tentukan file input yang akan digunakan
            if date in st.session_state.additional_files:
                current_input_file = st.session_state.additional_files[date]
                current_input_df = pd.read_excel(current_input_file, sheet_name=0, header=1)
            else:
                current_input_df = input_data_df
            
            # Cari kolom tanggal
            date_col = None
            for col in template_df.columns:
                if str(date) in str(col):
                    date_col = col
                    break
            
            if date_col is None:
                continue
            
            template_df[date_col] = template_df[date_col].astype(object)
            
            # Statistik per tanggal
            matched_count = 0
            not_matched_count = 0
            empty_value_count = 0
            
            # Isi data untuk tanggal ini
            for index, row in template_df.iterrows():
                crew_id = str(row["Crew ID"]).strip()
                match = current_input_df[current_input_df["Crew ID"].astype(str).str.strip() == crew_id]
                
                if not match.empty:
                    val = match[date_col].iloc[0] if date_col in current_input_df.columns else None
                    if pd.notna(val):
                        template_df.at[index, date_col] = val
                        matched_count += 1
                    else:
                        template_df.at[index, date_col] = "-"
                        empty_value_count += 1
                else:
                    template_df.at[index, date_col] = "-"
                    not_matched_count += 1
            
            # Simpan statistik tanggal ini
            overall_stats['date_details'][date] = {
                'matched': matched_count,
                'not_matched': not_matched_count,
                'empty': empty_value_count,
                'file_used': current_input_file.name if date in st.session_state.additional_files else input_file.name
            }
            overall_stats['total_matched'] += matched_count
            overall_stats['total_not_matched'] += not_matched_count
            overall_stats['total_empty'] += empty_value_count
            overall_stats['dates_processed'] += 1
            
            # Update progress
            progress = 5 + int(((idx + 1) / total_dates) * 75)
            progress_bar.progress(progress)
        
        progress_bar.progress(80)
        status_text.text("📝 Membuat file Excel...")
        
        # Buat workbook dengan formatting
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Header
        ws.append(["OPS - REPORT"])
        if apply_formatting:
            header_cell = ws['A1']
            header_cell.font = Font(bold=True, size=14, color="FFFFFF")
            header_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data
        for r_idx, r in enumerate(dataframe_to_rows(template_df, index=False, header=True)):
            ws.append(r)
            if apply_formatting and r_idx == 0:  # Header row
                for cell in ws[ws.max_row]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
        
        # Auto-adjust column width
        if apply_formatting:
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        progress_bar.progress(100)
        status_text.text("✅ Selesai!")
        
        end_time = time.time()
        process_time = end_time - start_time
        
        # Tampilkan hasil
        st.success(f"✅ Data berhasil diproses untuk {overall_stats['dates_processed']} tanggal dalam {process_time:.2f} detik!")
        
        # Statistik hasil keseluruhan
        if show_stats:
            st.subheader("📊 Statistik Hasil Proses")
            
            # Ringkasan keseluruhan
            col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
            with col_stat1:
                st.metric("Total Tanggal", overall_stats['dates_processed'])
            with col_stat2:
                avg_match = overall_stats['total_matched'] / overall_stats['dates_processed'] if overall_stats['dates_processed'] > 0 else 0
                st.metric("Rata-rata Match/Tanggal", f"{avg_match:.1f}")
            with col_stat3:
                st.metric("Total Data Cocok", overall_stats['total_matched'])
            with col_stat4:
                st.metric("Total Tidak Cocok", overall_stats['total_not_matched'])
            
            # Detail per tanggal
            st.divider()
            st.subheader("📅 Detail per Tanggal")
            
            detail_data = []
            for date in sorted(overall_stats['date_details'].keys()):
                detail = overall_stats['date_details'][date]
                detail_data.append({
                    'Tanggal': date,
                    'Data Cocok': detail['matched'],
                    'Tidak Cocok': detail['not_matched'],
                    'Nilai Kosong': detail['empty'],
                    'File Input': detail['file_used']
                })
            
            detail_df = pd.DataFrame(detail_data)
            st.dataframe(detail_df, use_container_width=True, hide_index=True)
        
        # Preview hasil
        st.subheader("📊 Preview Hasil Akhir")
        st.dataframe(template_df, use_container_width=True)
        
        # Tombol download
        col_dl1, col_dl2, col_dl3 = st.columns([1, 2, 1])
        with col_dl2:
            st.download_button(
                label="⬇️ Download Hasil sebagai Excel",
                data=output,
                file_name=output_filename + ".xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )
        
        # Log aktivitas
        with st.expander("📝 Log Aktivitas"):
            log_text = f"""
Waktu Proses: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Durasi: {process_time:.2f} detik
File Template: {template_file.name}
File Input Utama: {input_file.name}
Tanggal Diproses: {', '.join(map(str, selected_dates))}
Total Tanggal: {overall_stats['dates_processed']}
Total Baris: {overall_stats['total_rows']}
Total Data Cocok: {overall_stats['total_matched']}
Total Tidak Cocok: {overall_stats['total_not_matched']}
Total Nilai Kosong: {overall_stats['total_empty']}

Detail File Tambahan:
"""
            if st.session_state.additional_files:
                for date, file in st.session_state.additional_files.items():
                    log_text += f"  - Tanggal {date}: {file.name}\n"
            else:
                log_text += "  - Tidak ada file tambahan\n"
            
            st.code(log_text)
        
        time.sleep(0.5)
        progress_bar.empty()
        status_text.empty()
        
    except Exception as e:
        progress_bar.empty()
        status_text.empty()
        st.error(f"❌ Terjadi kesalahan saat memproses: {str(e)}")
        with st.expander("🔍 Detail Error"):
            st.exception(e)

elif template_df is None or input_data_df is None:
    st.warning("⚠️ Silakan upload kedua file terlebih dahulu")
elif not selected_dates:
    st.warning("⚠️ Silakan pilih tanggal yang ingin diproses")
elif len(validation_errors) > 0:
    st.error("❌ Perbaiki error validasi terlebih dahulu sebelum memproses")

# Footer
st.divider()
st.caption("💡 Tips: Pastikan kolom 'Crew ID' ada di semua file dan kolom tanggal sesuai dengan tanggal yang dipilih")
